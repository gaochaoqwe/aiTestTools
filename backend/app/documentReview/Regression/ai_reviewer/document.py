"""
文档生成模块 - 生成不同格式的审查文档
"""
import os
import uuid
import json
import logging
from datetime import datetime

def generate_review_document(review_results, output_folder):
    """
    生成需求审查文档，并返回相关的元数据
    
    参数:
        review_results: list, 需求审查结果列表
        output_folder: str, 输出文件夹
    
    返回:
        dict: 包含生成文档信息的字典
    """
    # 生成一个唯一的文件名
    doc_id = str(uuid.uuid4())
    
    # 尝试生成Excel格式，如果失败则回退到文本格式
    try:
        # 先尝试Excel格式
        file_name = f"requirement_review_{doc_id}.xlsx"
        output_path = os.path.join(output_folder, file_name)
        
        # 尝试导入必要的库
        import_success = True
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
        except ImportError:
            import_success = False
            logging.warning("无法导入pandas或openpyxl，将使用JSON格式")
        
        if import_success:
            # 使用Excel格式
            return generate_excel_document(review_results, output_path)
        else:
            # 回退到JSON格式
            return generate_json_document(review_results, output_folder, doc_id)
    
    except Exception as e:
        logging.error(f"生成Excel格式文档失败: {e}")
        # 回退到JSON格式
        return generate_json_document(review_results, output_folder, doc_id)

def generate_json_document(review_results, output_folder, doc_id=None):
    """
    生成JSON格式的需求审查文档
    
    参数:
        review_results: list, 需求审查结果列表
        output_folder: str, 输出文件夹
        doc_id: str, 可选的文档ID
        
    返回:
        dict: 包含生成文档信息的字典
    """
    if doc_id is None:
        doc_id = str(uuid.uuid4())
    
    file_name = f"requirement_review_{doc_id}.json"
    output_path = os.path.join(output_folder, file_name)
    
    # 确保输出目录存在
    os.makedirs(output_folder, exist_ok=True)
    
    # 生成JSON数据
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    output_data = {
        "meta": {
            "generated_time": timestamp,
            "document_id": doc_id,
            "format": "json"
        },
        "review_results": review_results
    }
    
    # 写入文件
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(output_data, f, ensure_ascii=False, indent=2)
    
    logging.info(f"已生成JSON格式审查文档: {output_path}")
    
    return {
        "document_id": doc_id,
        "file_path": output_path,
        "file_name": file_name,
        "format": "json",
        "generated_time": timestamp
    }

def generate_excel_document(review_results, output_path):
    """
    生成Excel格式的需求审查文档
    
    参数:
        review_results: list, 需求审查结果列表
        output_path: str, 输出文件路径
        
    返回:
        dict: 包含生成文档信息的字典
    """
    import pandas as pd
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
    
    # 确保输出目录存在
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    # 提取数据，创建表格数据结构
    rows = []
    for req_review in review_results:
        req_name = req_review.get("name", "未命名需求")
        req_chapter = req_review.get("chapter", "")
        review_result = req_review.get("review_result", {})
        
        # 获取审查问题列表
        problems = review_result.get("requirements_review", [])
        score = review_result.get("score", 0)
        summary = review_result.get("summary", "")
        
        # 如果没有问题，创建一个空行
        if not problems:
            rows.append({
                "需求名称": req_name,
                "章节": req_chapter,
                "问题标题": "无问题",
                "需求描述": "",
                "问题描述": "",
                "问题位置": "",
                "影响分析": "",
                "分数": score,
                "总结": summary
            })
        else:
            # 添加每个问题
            for i, problem in enumerate(problems):
                rows.append({
                    "需求名称": req_name if i == 0 else "",
                    "章节": req_chapter if i == 0 else "",
                    "问题标题": problem.get("problem_title", ""),
                    "需求描述": problem.get("requirement_description", ""),
                    "问题描述": problem.get("problem_description", ""),
                    "问题位置": problem.get("problem_location", ""),
                    "影响分析": problem.get("impact_analysis", ""),
                    "分数": score if i == 0 else "",
                    "总结": summary if i == 0 else ""
                })
    
    # 创建DataFrame
    df = pd.DataFrame(rows)
    
    # 保存为Excel
    df.to_excel(output_path, index=False)
    
    # 美化Excel
    wb = load_workbook(output_path)
    ws = wb.active
    
    # 设置列宽
    ws.column_dimensions['A'].width = 20  # 需求名称
    ws.column_dimensions['B'].width = 10  # 章节
    ws.column_dimensions['C'].width = 25  # 问题标题
    ws.column_dimensions['D'].width = 25  # 需求描述
    ws.column_dimensions['E'].width = 35  # 问题描述
    ws.column_dimensions['F'].width = 20  # 问题位置
    ws.column_dimensions['G'].width = 35  # 影响分析
    ws.column_dimensions['H'].width = 8   # 分数
    ws.column_dimensions['I'].width = 30  # 总结
    
    # 设置自动换行
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical='top')
    
    # 设置标题行格式
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # 保存修改后的Excel
    wb.save(output_path)
    
    # 生成文档ID
    doc_id = os.path.basename(output_path).split('.')[0].replace('requirement_review_', '')
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    
    logging.info(f"已生成Excel格式审查文档: {output_path}")
    
    return {
        "document_id": doc_id,
        "file_path": output_path,
        "file_name": os.path.basename(output_path),
        "format": "excel",
        "generated_time": timestamp
    }

def generate_review_doc(review_results, output_path, format_type="json"):
    """
    生成需求审查文档
    
    参数:
        review_results: list, 需求审查结果列表
        output_path: str, 输出文件路径
        format_type: str, 文档格式，支持json、text和excel
    
    返回:
        str: 生成的文件路径
    """
    logging.info(f"开始生成需求审查文档: {output_path}，格式: {format_type}")
    
    # 确保输出目录存在
    os.makedirs(os.path.dirname(output_path), exist_ok=True)
    
    if format_type == "json":
        # 生成JSON格式
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(review_results, f, ensure_ascii=False, indent=2)
        
        logging.info(f"已生成JSON格式审查文档: {output_path}")
        return output_path
    
    elif format_type == "text":
        # 生成文本格式
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("需求审查报告\n")
            f.write("=" * 50 + "\n\n")
            
            for req_review in review_results:
                req_name = req_review.get("name", "未命名需求")
                req_chapter = req_review.get("chapter", "")
                review_result = req_review.get("review_result", {})
                
                f.write(f"需求名称: {req_name}\n")
                if req_chapter:
                    f.write(f"章节: {req_chapter}\n")
                
                problems = review_result.get("requirements_review", [])
                score = review_result.get("score", 0)
                summary = review_result.get("summary", "")
                
                f.write(f"审查得分: {score}\n")
                f.write(f"总结: {summary}\n\n")
                
                if problems:
                    f.write("发现的问题:\n")
                    for i, problem in enumerate(problems):
                        f.write(f"问题 {i+1}: {problem.get('problem_title', '')}\n")
                        f.write(f"  需求描述: {problem.get('requirement_description', '')}\n")
                        f.write(f"  问题描述: {problem.get('problem_description', '')}\n")
                        f.write(f"  问题位置: {problem.get('problem_location', '')}\n")
                        f.write(f"  影响分析: {problem.get('impact_analysis', '')}\n\n")
                else:
                    f.write("未发现问题\n\n")
                
                f.write("-" * 50 + "\n\n")
        
        logging.info(f"已生成文本格式审查文档: {output_path}")
        return output_path
    
    elif format_type == "excel":
        # 尝试生成Excel格式
        try:
            import pandas as pd
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
            
            # 提取数据，创建表格数据结构
            rows = []
            for req_review in review_results:
                req_name = req_review.get("name", "未命名需求")
                req_chapter = req_review.get("chapter", "")
                review_result = req_review.get("review_result", {})
                
                # 获取审查问题列表
                problems = review_result.get("requirements_review", [])
                score = review_result.get("score", 0)
                summary = review_result.get("summary", "")
                
                # 如果没有问题，创建一个空行
                if not problems:
                    rows.append({
                        "需求名称": req_name,
                        "章节": req_chapter,
                        "问题标题": "无问题",
                        "需求描述": "",
                        "问题描述": "",
                        "问题位置": "",
                        "影响分析": "",
                        "分数": score,
                        "总结": summary
                    })
                else:
                    # 添加每个问题
                    for i, problem in enumerate(problems):
                        rows.append({
                            "需求名称": req_name if i == 0 else "",
                            "章节": req_chapter if i == 0 else "",
                            "问题标题": problem.get("problem_title", ""),
                            "需求描述": problem.get("requirement_description", ""),
                            "问题描述": problem.get("problem_description", ""),
                            "问题位置": problem.get("problem_location", ""),
                            "影响分析": problem.get("impact_analysis", ""),
                            "分数": score if i == 0 else "",
                            "总结": summary if i == 0 else ""
                        })
            
            # 创建DataFrame
            df = pd.DataFrame(rows)
            
            # 保存为Excel
            df.to_excel(output_path, index=False)
            
            # 美化Excel
            wb = load_workbook(output_path)
            ws = wb.active
            
            # 设置列宽
            ws.column_dimensions['A'].width = 20  # 需求名称
            ws.column_dimensions['B'].width = 10  # 章节
            ws.column_dimensions['C'].width = 25  # 问题标题
            ws.column_dimensions['D'].width = 25  # 需求描述
            ws.column_dimensions['E'].width = 35  # 问题描述
            ws.column_dimensions['F'].width = 20  # 问题位置
            ws.column_dimensions['G'].width = 35  # 影响分析
            ws.column_dimensions['H'].width = 8   # 分数
            ws.column_dimensions['I'].width = 30  # 总结
            
            # 设置自动换行
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
            
            # 设置标题行格式
            header_font = Font(bold=True, size=12)
            header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), 
                right=Side(style='thin'), 
                top=Side(style='thin'), 
                bottom=Side(style='thin')
            )
            
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            
            # 保存修改后的Excel
            wb.save(output_path)
            
            logging.info(f"已生成Excel格式审查文档: {output_path}")
            return output_path
        
        except ImportError:
            logging.warning("无法导入pandas或openpyxl，将使用JSON格式")
            # 回退到JSON格式
            output_path = output_path.replace('.xlsx', '.json')
            return generate_review_doc(review_results, output_path, "json")
        
        except Exception as e:
            logging.error(f"生成Excel文档时出错: {str(e)}")
            # 回退到JSON格式
            output_path = output_path.replace('.xlsx', '.json')
            return generate_review_doc(review_results, output_path, "json")
    
    else:
        # 不支持的格式，回退到JSON
        logging.warning(f"不支持的文档格式: {format_type}，将使用JSON格式")
        output_path = output_path.replace(f'.{format_type}', '.json')
        return generate_review_doc(review_results, output_path, "json")
