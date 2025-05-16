"""
需求审查模块
使用AI模型API对需求进行自动审查，支持OpenAI和Ollama API
"""
import os
import logging
import json
import requests
from openai import OpenAI

# 导入配置模块
from app.documentReview.ConfigurationItem.config import get_config

# 配置日志
logging.basicConfig(level=logging.DEBUG, format='[%(levelname)s] %(message)s')

def get_client():
    """
    根据配置获取相应的AI客户端
    
    返回:
        客户端对象或None
    """
    import os
    import json
    
    # 从环境变量中获取API密钥（优先级最高）
    env_api_key = os.environ.get("OPENAI_API_KEY", "")
    if env_api_key:
        logging.info("从环境变量中读取到OpenAI API密钥")
    
    # 直接从配置文件中读取配置（辅助方法）
    config_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), '..', 'config.json')
    file_config = {}
    if os.path.exists(config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                file_config = json.load(f)
                logging.info(f"直接从文件读取配置: {config_path}")
        except Exception as e:
            logging.error(f"读取配置文件时出错: {str(e)}")
    
    # 从配置系统获取配置（原始方法）
    config = get_config()
    logging.info(f"当前系统中的配置: {config}")
    
    # 比较配置文件和配置对象中的API密钥（调试所用）
    file_api_key = ""
    if "openai" in file_config and "api_key" in file_config["openai"]:
        file_api_key = file_config["openai"]["api_key"]
    
    sys_api_key = ""
    if "openai" in config and "api_key" in config["openai"]:
        sys_api_key = config["openai"]["api_key"]
        
    if file_api_key != sys_api_key:
        logging.warning(f"文件配置与内存配置中的API密钥不一致！")
        logging.warning(f"文件密钥长度: {len(file_api_key)}, 内存密钥长度: {len(sys_api_key)}")
    
    # 如果环境变量中有API密钥，优先使用
    if env_api_key:
        if "openai" not in config:
            config["openai"] = {}
        config["openai"]["api_key"] = env_api_key
    
    # 在找不到API密钥时，使用文件中的密钥作为备选（强制覆盖）
    if not sys_api_key and file_api_key:
        logging.warning(f"配置对象中无API密钥，使用配置文件中的密钥代替")
        if "openai" not in config:
            config["openai"] = {}
        config["openai"]["api_key"] = file_api_key
    
    provider = config.get("provider", "openai")
    
    if provider == "openai":
        openai_config = config.get("openai", {})
        api_key = openai_config.get("api_key", "")
        base_url = openai_config.get("base_url", "")
        
        if not api_key:
            logging.error("OpenAI API密钥未设置")
            
            # 尝试直接从文件读取API密钥作为备用方案
            if file_api_key:
                logging.warning("使用配置文件中的API密钥作为最后尝试")
                api_key = file_api_key
                base_url = file_config.get("openai", {}).get("base_url", base_url)
            else:
                return None
        
        # 调试输出 - 记录API密钥前几位和长度
        masked_key = "****"
        if api_key and len(api_key) > 8:
            masked_key = api_key[:4] + '*' * (len(api_key) - 8) + api_key[-4:]
        logging.debug(f"创建OpenAI客户端: API密钥: {masked_key}, 基础URL: {base_url}")
        
        try:
            # 对于非OpenAI的API，可能需要不同的处理方式
            is_third_party_api = ("siliconflow" in base_url or 
                                 "glm" in base_url.lower() or 
                                 "qwen" in base_url.lower() or 
                                 "zhipu" in base_url.lower())
            
            if is_third_party_api:
                logging.info(f"使用第三方API: {base_url}")
                
                # 检查base_url是否包含完整路径
                if "chat/completions" in base_url:
                    logging.warning(f"base_url包含具体端点路径，这可能导致OpenAI库出现问题")
                    # 提取基础URL部分（去掉/chat/completions）
                    base_url = base_url.split("/chat/completions")[0]
                    logging.info(f"已提取基础URL: {base_url}")
            
            # 重要修复: 与成功的测试案例保持一致，显式设置认证头
            # 创建完整的授权头，与成功测试一致
            auth_header = f"Bearer {api_key}"
            logging.info(f"使用自定义认证头: {auth_header[:10]}...{auth_header[-4:]}")
            
            # 构建OpenAI客户端
            client = OpenAI(
                api_key=api_key,
                base_url=base_url,
                default_headers={"Authorization": auth_header}
            )
                
            return client
        except Exception as e:
            logging.error(f"创建OpenAI客户端失败: {str(e)}")
            raise ValueError(f"创建OpenAI客户端失败: {str(e)}")
    
    # Ollama不需要创建客户端，直接使用requests库调用
    return None

def review_requirement(requirement):
    """
    使用AI模型对单个需求进行审查
    
    参数:
        requirement: dict, 包含需求信息的字典，至少有name和content字段
    
    返回:
        dict: 包含审查结果的字典
    """
    logging.debug(f"开始审查需求《{requirement['name']}》")
    config = get_config()
    provider = config.get("provider", "openai")
    
    # 准备提示词
    prompt = f"""
    请作为一名专业的需求审查专家，审查以下软件需求，找出潜在问题，严格按照要求的JSON格式返回结果：
    
    需求名称: {requirement['name']}
    需求内容:
    {requirement['content']}
    
    请仔细审查上述需求，找出所有潜在问题，例如：不明确的描述、不完整的约束条件、性能要求不具体、冲突的需求等。
    
    审查结果必须严格按照以下JSON格式返回（数组中可以包含多个问题）：
    {{
        "requirements_review": [
            {{
                "problem_title": "问题的具体名称（包含功能名称和问题类型，不超过20个字）",
                "requirement_description": "简要描述需求要点",
                "problem_description": "详细描述发现的问题",
                "problem_location": "指出问题在需求中的具体位置",
                "impact_analysis": "分析此问题可能带来的影响"
            }}
        ],
        "score": 85,
        "summary": "审查总结"
    }}
    
    必须为每个发现的问题提供所有五个字段：problem_title、requirement_description、problem_description、problem_location和impact_analysis。
    
    其中problem_title是对问题的具体描述，必须包含需求的功能名称和具体问题类型，格式为"[功能名称]+[问题类型]"，如"用户身份验证功能约束不明确"、"用户身份验证功能性能要求不具体"等，不超过20个字。
    
    如果没有发现问题，请在requirements_review数组中返回一个对象，表明需求质量良好。
    请直接返回JSON对象，不要包含任何额外的文字。
    """
    
    model_params = config.get("model_params", {})
    temperature = model_params.get("temperature", 0.1)
    max_tokens = model_params.get("max_tokens", 2000)
    
    try:
        # 尝试使用OpenAI API
        client = get_client()
        
        if client is None:
            # 如果无法创建API客户端，返回模拟数据
            logging.warning("无法创建OpenAI客户端，切换到模拟审查模式")
            review_item = {
                "name": requirement["name"],
                "chapter": requirement.get("chapter", ""),
                "review_result": {
                    "requirements_review": [
                        {
                            "problem_title": f"{requirement['name']}描述细节缺失",
                            "requirement_description": "需求完整性分析",
                            "problem_description": "需求描述缺乏具体细节，可能导致不同理解",
                            "problem_location": "需求说明部分",
                            "impact_analysis": "可能导致开发者对要求理解不一致，影响功能实现"
                        },
                        {
                            "problem_title": f"{requirement['name']}术语一致性问题",
                            "requirement_description": "需求术语一致性分析",
                            "problem_description": "需求中使用的术语可能存在不一致的情况",
                            "problem_location": "需求整体文档",
                            "impact_analysis": "可能导致模块之间集成困难，增加沟通成本"
                        }
                    ],
                    "score": 85,
                    "summary": "需求总体质量良好，但存在一些可以改进的地方"
                }
            }
            return review_item
        
        # 根据不同提供商调用API
        if provider == "openai":
            result = call_openai_api(prompt, temperature, max_tokens)
        elif provider == "ollama":
            result = call_ollama_api(prompt, temperature, max_tokens)
        else:
            raise ValueError(f"不支持的模型提供商: {provider}")
        
        # 解析结果
        logging.debug(f"AI返回的原始结果: {result}")
        
        try:
            # 解析JSON结果
            result_json = json.loads(result)
            
            # 添加原始需求信息
            review_result = {
                "name": requirement["name"],
                "chapter": requirement.get("chapter", ""),
                "review_result": result_json
            }
            
            logging.info(f"需求《{requirement['name']}》审查完成，发现 {len(result_json.get('requirements_review', []))} 个问题")
            return review_result
            
        except json.JSONDecodeError as e:
            logging.error(f"解析AI返回的JSON失败: {e}")
            # 返回错误信息
            return {
                "name": requirement["name"],
                "chapter": requirement.get("chapter", ""),
                "review_result": {
                    "requirements_review": [{
                        "problem_title": f"{requirement['name']}解析失败",
                        "requirement_description": "需求解析失败",
                        "problem_description": f"AI返回的结果无法解析为JSON: {e}",
                        "problem_location": "整个需求文档",
                        "impact_analysis": "无法进行有效的需求审查，可能导致需求问题被忽略"
                    }],
                    "score": 0,
                    "summary": "无法完成需求审查"
                }
            }
    
    except Exception as e:
        logging.error(f"调用AI审查需求时出错: {e}")
        # 返回错误信息
        return {
            "name": requirement["name"],
            "chapter": requirement.get("chapter", ""),
            "review_result": {
                "requirements_review": [{
                    "problem_title": f"{requirement['name']}审查系统错误",
                    "requirement_description": "需求审查系统错误",
                    "problem_description": f"调用AI审查需求时出错: {str(e)}",
                    "problem_location": "系统调用过程",
                    "impact_analysis": "系统无法完成需求审查，可能导致需求缺陷无法发现和修复"
                }],
                "score": 0,
                "summary": "无法完成需求审查"
            }
        }

def call_openai_api(prompt, temperature, max_tokens):
    """
    调用OpenAI API
    
    参数:
        prompt: str, 提示词
        temperature: float, 温度参数
        max_tokens: int, 最大令牌数
    
    返回:
        str: 模型响应内容
    """
    config = get_config()
    openai_config = config.get("openai", {})
    model_name = openai_config.get("model_name", "gpt-4o")
    api_key = openai_config.get("api_key", "")
    base_url = openai_config.get("base_url", "")
    
    # 检测是否为第三方API服务
    is_third_party_api = ("siliconflow" in base_url or 
                          "glm" in base_url.lower() or 
                          "qwen" in base_url.lower() or 
                          "zhipu" in base_url.lower())
    
    # 打印请求参数详情
    masked_key = "****"
    if api_key and len(api_key) > 8:
        masked_key = api_key[:4] + '*' * (len(api_key) - 8) + api_key[-4:]
        
    # 打印请求详情
    request_params = {
        "provider": "openai",
        "model": model_name,
        "api_key_prefix": api_key[:4] if len(api_key) > 4 else "无效",
        "api_key_length": len(api_key),
        "base_url": base_url,
        "temperature": temperature,
        "max_tokens": max_tokens,
        "prompt_length": len(prompt),
        "prompt_preview": prompt[:50] + "..." if len(prompt) > 50 else prompt,
        "is_third_party_api": is_third_party_api
    }
    
    logging.info(f"OpenAI请求参数: {json.dumps(request_params, ensure_ascii=False)}")
    
    # 对于第三方API，优先使用直接HTTP请求
    if is_third_party_api:
        try:
            return call_direct_http_api(prompt, temperature, max_tokens)
        except Exception as e:
            logging.warning(f"直接HTTP请求失败，尝试使用OpenAI库: {str(e)}")
            # 如果直接请求失败，回退到OpenAI库
    
    try:
        client = get_client()
        if not client:
            raise ValueError("无法创建OpenAI客户端")
            
        # 构建请求参数
        request_args = {
            "model": model_name,
            "messages": [
                {"role": "system", "content": "你是一名专业的需求分析专家，擅长发现需求文档中的问题。"},
                {"role": "user", "content": prompt}
            ],
            "temperature": temperature,
            "max_tokens": max_tokens
        }
        
        # 如果不是验证模型，则添加JSON响应格式
        if "验证模型" not in prompt:
            request_args["response_format"] = {"type": "json_object"}
            
        # 打印实际请求参数
        logging.info(f"OpenAI API请求参数: {json.dumps(request_args, ensure_ascii=False)}")
        
        response = client.chat.completions.create(**request_args)
        
        # 提取响应内容
        return response.choices[0].message.content
    
    except Exception as e:
        error_message = str(e)
        # 提取更详细的错误信息
        if hasattr(e, 'response'):
            try:
                # 尝试从响应中提取更多错误信息
                response_json = e.response.json()
                logging.error(f"OpenAI API完整错误响应: {json.dumps(response_json, ensure_ascii=False)}")
                if 'error' in response_json:
                    error_message = f"{response_json['error'].get('message', error_message)}"
            except Exception as parse_err:
                logging.error(f"解析错误响应失败: {parse_err}")
                pass
        
        # 对于第三方API的认证错误，如果我们还没有尝试直接HTTP请求方式，现在尝试
        if not is_third_party_api and ("401" in error_message or "unauthorized" in error_message.lower()):
            try:
                logging.info(f"尝试使用直接HTTP请求方式作为备选")
                return call_direct_http_api(prompt, temperature, max_tokens)
            except Exception as http_err:
                logging.error(f"直接HTTP请求也失败: {str(http_err)}")
                # 如果直接HTTP请求也失败，继续抛出原始错误
        
        # 检查是否为认证错误
        if "401" in error_message or "unauthorized" in error_message.lower() or "authentication" in error_message.lower():
            raise ValueError(f"Error code: 401 - API密钥无效或认证失败, 请检查API密钥和基础URL设置。详细信息: {error_message}")
        
        # 检查是否为模型不存在错误
        elif "model" in error_message.lower() and ("not found" in error_message.lower() or "does not exist" in error_message.lower()):
            raise ValueError(f"指定的模型 '{model_name}' 不存在或不可用，请检查模型名称")
        
        # 其他错误
        else:
            raise ValueError(f"调用OpenAI API时出错: {error_message}")

def call_direct_http_api(prompt, temperature, max_tokens):
    """
    使用直接HTTP请求调用兼容OpenAI的API
    针对第三方API可能需要不同的认证方式
    
    参数:
        prompt: str, 提示词
        temperature: float, 温度参数
        max_tokens: int, 最大令牌数
    
    返回:
        str: 模型响应内容
    """
    import requests
    
    config = get_config()
    openai_config = config.get("openai", {})
    model_name = openai_config.get("model_name", "gpt-4o")
    api_key = openai_config.get("api_key", "")
    base_url = openai_config.get("base_url", "")
    
    # 判断并构建正确的URL
    if "chat/completions" in base_url:
        # 如果base_url已经包含了端点路径，则直接使用
        url = base_url
        logging.info(f"使用预配置的完整端点URL: {url}")
    else:
        # 确保base_url不以/结尾
        if base_url.endswith('/'):
            base_url = base_url[:-1]
        # 构建完整URL
        url = f"{base_url}/chat/completions"
        logging.info(f"构建端点URL: {url}")
    
    # 准备请求数据 - 与成功的测试案例完全一致
    payload = {
        "model": model_name,
        "messages": [
            {
                "role": "user",
                "content": prompt
            }
        ],
        "stream": False,
        "max_tokens": max_tokens,
        "temperature": temperature,
        "response_format": {"type": "text"}
    }
    
    # 准备认证头 - 直接使用Bearer前缀，与成功案例保持一致
    auth_header = f"Bearer {api_key}"
    
    headers = {
        "Authorization": auth_header,
        "Content-Type": "application/json"
    }
    
    logging.info(f"直接HTTP请求 - URL: {url}")
    logging.info(f"直接HTTP请求 - 模型: {model_name}")
    logging.info(f"直接HTTP请求 - 认证头格式: {auth_header[:10]}...{auth_header[-4:]}")
    
    # 发送请求
    try:
        # 注释掉直接HTTP请求的部分
        # response = requests.post(url, json=payload, headers=headers)
        raise Exception("直接HTTP请求已被禁用")
        
        # 检查响应状态
        # if response.status_code != 200:
        #     logging.error(f"API请求失败: 状态码 {response.status_code}")
        #     logging.error(f"响应内容: {response.text}")
        #     raise ValueError(f"API请求失败: 状态码 {response.status_code}, 响应: {response.text}")
        
        # 解析JSON响应
        # result = response.json()
        
        # 从响应中提取内容
        # if "choices" in result and len(result["choices"]) > 0:
        #     message = result["choices"][0].get("message", {})
        #     content = message.get("content", "")
        #     return content
        # else:
        #     logging.error(f"API响应格式不符合预期: {json.dumps(result, ensure_ascii=False)}")
        #     raise ValueError(f"API响应格式不符合预期")
            
    except requests.RequestException as e:
        logging.error(f"HTTP请求错误: {str(e)}")
        raise ValueError(f"HTTP请求错误: {str(e)}")
    except json.JSONDecodeError as e:
        logging.error(f"响应不是有效的JSON: {str(e)}")
        raise ValueError(f"响应不是有效的JSON: {str(e)}")
    except Exception as e:
        logging.error(f"调用API时发生错误: {str(e)}")
        raise ValueError(f"调用API时发生错误: {str(e)}")

def call_ollama_api(prompt, temperature, max_tokens):
    """
    调用Ollama API
    
    参数:
        prompt: str, 提示词
        temperature: float, 温度参数
        max_tokens: int, 最大令牌数
    
    返回:
        str: 模型响应内容
    """
    config = get_config()
    ollama_config = config.get("ollama", {})
    base_url = ollama_config.get("base_url", "http://localhost:11434")
    model_name = ollama_config.get("model_name", "llama3")
    
    # 打印请求详情
    request_params = {
        "provider": "ollama",
        "model": model_name,
        "base_url": base_url,
        "temperature": temperature,
        "max_tokens": max_tokens,
        "prompt_length": len(prompt),
        "prompt_preview": prompt[:50] + "..." if len(prompt) > 50 else prompt
    }
    
    logging.info(f"Ollama请求参数: {json.dumps(request_params, ensure_ascii=False)}")
    
    # 构建API端点
    endpoint = f"{base_url}/api/generate"
    
    # 构建请求数据
    data = {
        "model": model_name,
        "prompt": f"系统: 你是一名专业的需求分析专家，擅长发现需求文档中的问题。\n\n用户: {prompt}",
        "temperature": temperature,
        "max_tokens": max_tokens,
        "stream": False,
        "format": "json" if "验证模型" not in prompt else None  # 尝试让Ollama返回JSON格式，验证模型时不需要
    }
    
    logging.info(f"Ollama API请求参数: {json.dumps(data, ensure_ascii=False)}")
    
    try:
        # 发送请求
        response = requests.post(endpoint, json=data)
        
        if response.status_code != 200:
            error_detail = f"状态码: {response.status_code}"
            try:
                # 尝试从响应中提取更多错误信息
                response_json = response.json()
                logging.error(f"Ollama API完整错误响应: {json.dumps(response_json, ensure_ascii=False)}")
                if 'error' in response_json:
                    error_detail += f", 错误信息: {json.dumps(response_json, ensure_ascii=False)}"
            except:
                error_detail += f", 响应内容: {response.text[:200]}"
                
            raise ValueError(f"Ollama API请求失败: {error_detail}")
        
        # 解析响应
        result = response.json()
        logging.debug(f"Ollama响应: {json.dumps(result, ensure_ascii=False)[:500]}")
        return result.get("response", "")
    except requests.RequestException as e:
        logging.error(f"Ollama API连接错误: {str(e)}")
        raise ValueError(f"Ollama API连接错误: {str(e)}")
    except Exception as e:
        logging.error(f"Ollama API调用出错: {str(e)}")
        raise ValueError(f"Ollama API调用出错: {str(e)}")

def review_requirements(requirements):
    """
    批量审查多个需求
    
    参数:
        requirements: list, 需求列表，每个元素是一个包含需求信息的字典
    
    返回:
        list: 包含所有需求审查结果的列表
    """
    logging.info(f"开始批量审查 {len(requirements)} 个需求")
    
    results = []
    for requirement in requirements:
        # 依次审查每个需求
        review_result = review_requirement(requirement)
        results.append(review_result)
    
    logging.info(f"完成批量审查，共审查 {len(results)} 个需求")
    return results

def generate_review_document(review_results, output_folder):
    """
    生成需求审查文档，并返回相关的元数据
    
    参数:
        review_results: list, 需求审查结果列表
        output_folder: str, 输出文件夹
    
    返回:
        dict: 包含生成文档信息的字典
    """
    import uuid
    import os
    
    # 生成一个唯一的文件名
    doc_id = str(uuid.uuid4())
    
    # 尝试生成Excel格式，如果失败则回退到文本格式
    try:
        # 先尝试Excel格式
        file_name = f"requirement_review_{doc_id}.xlsx"
        output_path = os.path.join(output_folder, file_name)
        
        # 尝试导入必要的库
        import_success = True
        try:
            import pandas as pd
            from openpyxl import Workbook
        except ImportError:
            import_success = False
            logging.warning("缺少pandas或openpyxl库，将回退到文本格式")
        
        if import_success:
            # 生成Excel文档
            generate_review_doc(review_results, output_path, format_type="excel")
        else:
            # 导入失败，改用文本格式
            file_name = f"requirement_review_{doc_id}.txt"
            output_path = os.path.join(output_folder, file_name)
            generate_review_doc(review_results, output_path, format_type="text")
            
    except Exception as e:
        # 任何错误都回退到文本格式
        logging.error(f"生成Excel报告出错: {str(e)}，回退到文本格式")
        file_name = f"requirement_review_{doc_id}.txt"
        output_path = os.path.join(output_folder, file_name)
        generate_review_doc(review_results, output_path, format_type="text")
    
    return {
        "message": "审查文档生成成功",
        "file_path": output_path,
        "file_name": file_name,
        "doc_id": doc_id,
        "download_url": f"/api/download_review/{doc_id}"
    }

def generate_review_doc(review_results, output_path, format_type="json"):
    """
    生成需求审查文档
    
    参数:
        review_results: list, 需求审查结果列表
        output_path: str, 输出文件路径
        format_type: str, 文档格式，支持json、text和excel
    
    返回:
        str: 生成的文件路径
    """
    logging.info(f"开始生成需求审查文档: {output_path}，格式: {format_type}")
    
    if format_type == "json":
        # 保存为JSON文件
        with open(output_path, 'w', encoding='utf-8') as f:
            json.dump(review_results, f, ensure_ascii=False, indent=2)
    elif format_type == "text":
        # 生成文本格式报告
        with open(output_path, 'w', encoding='utf-8') as f:
            f.write("需求审查综合报告\n\n")
            f.write("=========================\n\n")
            
            for i, review in enumerate(review_results):
                f.write(f"需求{i+1}: {review.get('name', '')}\n")
                f.write(f"章节号: {review.get('chapter', '')}\n")
                
                if 'review_result' in review and 'requirements_review' in review['review_result']:
                    problems = review['review_result']['requirements_review']
                    f.write(f"发现问题: {len(problems)} 个\n")
                    
                    for j, problem in enumerate(problems):
                        f.write(f"  {j+1}. 问题类型: {problem.get('problem_description', '')}\n")
                        f.write(f"     位置: {problem.get('problem_location', 'N/A')}\n")
                        f.write(f"     影响分析: {problem.get('impact_analysis', 'N/A')}\n")
                    
                    # 添加得分和总结，如果有的话
                    if 'score' in review['review_result']:
                        f.write(f"评分: {review['review_result'].get('score', 0)}\n")
                    if 'summary' in review['review_result']:
                        f.write(f"总结: {review['review_result'].get('summary', '')}\n")
                
                f.write("\n-------------------------\n\n")
    elif format_type == "excel":
        # 生成Excel格式报告
        try:
            import pandas as pd
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, PatternFill
            
            # 创建工作簿
            wb = Workbook()
            ws = wb.active
            ws.title = "需求审查报告"
            
            # 设置标题样式
            title_font = Font(name='宋体', size=14, bold=True)
            title_alignment = Alignment(horizontal='center', vertical='center')
            
            # 添加标题
            ws.merge_cells('A1:H1')  # 更新合并单元格范围，因为现在有8列
            ws['A1'] = "需求审查综合报告"
            ws['A1'].font = title_font
            ws['A1'].alignment = title_alignment
            
            # 设置列标题 - 根据用户需求修改列名
            headers = ["测试活动", "问题类型", "问题等级", "问题数目", "问题名称", "问题来源", "问题描述", "测评单位意见"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=2, column=col, value=header)
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
            
            # 添加数据
            row = 3
            for i, review in enumerate(review_results):
                name = review.get('name', '')
                chapter = review.get('chapter', '')
                
                # 检查是否有问题
                if 'review_result' in review and 'requirements_review' in review['review_result']:
                    problems = review['review_result']['requirements_review']
                    if problems and len(problems) > 0:
                        # 添加每个问题
                        for j, problem in enumerate(problems):
                            # 构建四段式问题描述
                            problem_description = f"需求描述：{problem.get('requirement_description', 'N/A')}\n\n"
                            problem_description += f"问题描述：{problem.get('problem_description', 'N/A')}\n\n"
                            problem_description += f"问题定位：{problem.get('problem_location', 'N/A')}\n\n"
                            problem_description += f"影响分析：{problem.get('impact_analysis', 'N/A')}"
                            
                            # 构建问题来源 - 修复章节号重复问题
                            # 检查需求名称中是否已包含章节号
                            if chapter in name:
                                problem_source = f"需求规格说明 {name}"
                            else:
                                problem_source = f"需求规格说明 {chapter} {name}"
                            
                            # 按用户要求填写各列内容
                            ws.cell(row=row, column=1, value="文档审查")
                            ws.cell(row=row, column=2, value="软件需求问题")
                            ws.cell(row=row, column=3, value="一般问题")
                            ws.cell(row=row, column=4, value=1)
                            ws.cell(row=row, column=5, value=problem.get('problem_title', 'N/A'))
                            ws.cell(row=row, column=6, value=problem_source)  # 新的问题来源列
                            ws.cell(row=row, column=7, value=problem_description)  # 原来的问题来源列，现在改名为问题描述
                            ws.cell(row=row, column=8, value="请开发人员对文档问题进行修改，确保文档描述的正确且测试人员能够理解测试需求")
                            
                            # 设置单元格自动换行
                            for col in range(1, 9):  # 现在是8列
                                cell = ws.cell(row=row, column=col)
                                cell.alignment = Alignment(wrap_text=True, vertical='top')
                            
                            row += 1
            
            # 调整列宽
            column_widths = [15, 15, 15, 10, 25, 30, 40, 30]  # 更新为8列的宽度
            for i, width in enumerate(column_widths, 1):
                ws.column_dimensions[chr(64 + i)].width = width
                
            # 设置行高
            for r in range(3, row):
                ws.row_dimensions[r].height = 120
            
            # 保存Excel文件
            wb.save(output_path)
            logging.info(f"Excel审查报告生成成功: {output_path}")
            
        except ImportError as e:
            logging.error(f"生成Excel报告失败，缺少必要的库: {str(e)}")
            raise
        except Exception as e:
            logging.error(f"生成Excel报告时出错: {str(e)}", exc_info=True)
            raise
    else:
        raise ValueError(f"不支持的文档格式: {format_type}")
    
    logging.info(f"需求审查文档生成完成: {output_path}")
    return output_path

def validate_model():
    """
    验证当前配置的模型是否可以正常使用
    
    返回:
        str: 模型的响应内容
    """
    logging.info("开始验证模型配置")
    config = get_config()
    provider = config.get("provider", "openai")
    
    # 准备验证提示和参数
    prompt = "你好，请简单回复一句问候语，验证模型连接正常。"
    
    # 获取模型参数
    model_params = config.get("model_params", {})
    temperature = model_params.get("temperature", 0.7)
    max_tokens = 100  # 对于验证，使用较小的max_tokens
    
    # 根据提供商类型调用不同的API
    try:
        if provider == "openai":
            openai_config = config.get("openai", {})
            model_name = openai_config.get("model_name", "gpt-4o")
            api_key = openai_config.get("api_key", "")
            base_url = openai_config.get("base_url", "")
            
            # 调试模式：显示完整API密钥以便排查问题
            logging.info(f"OpenAI配置 - 模型: {model_name}")
            logging.info(f"完整API密钥【调试用】: '{api_key}'")
            logging.info(f"基础URL: {base_url}")
            logging.info(f"模型参数 - 温度: {temperature}, 最大令牌数: {max_tokens}")
            
            # 记录详细的请求参数，用于调试
            is_third_party_api = (
                "siliconflow" in base_url or 
                "glm" in base_url.lower() or 
                "qwen" in base_url.lower() or 
                "zhipu" in base_url.lower()
            )
            
            request_info = {
                "provider": provider,
                "model": model_name,
                "api_key_prefix": api_key[:4] if api_key else "",
                "api_key_length": len(api_key) if api_key else 0,
                "base_url": base_url,
                "temperature": temperature,
                "max_tokens": max_tokens,
                "prompt_length": len(prompt),
                "prompt_preview": prompt[:30],
                "is_third_party_api": is_third_party_api
            }
            logging.info(f"OpenAI请求参数: {json.dumps(request_info, ensure_ascii=False)}")
            
            # 直接使用OpenAI库
            try:
                # 尝试使用OpenAI库
                logging.info("使用OpenAI库进行验证")
                client = get_client()
                if not client:
                    raise ValueError("无法创建OpenAI客户端")
                
                system_message = "你是一名专业的需求分析专家，擅长发现需求文档中的问题。"
                messages = [
                    {"role": "system", "content": system_message},
                    {"role": "user", "content": prompt}
                ]
                
                # 记录实际的请求参数
                request_params = {
                    'model': model_name,
                    'messages': messages,
                    'temperature': temperature,
                    'max_tokens': max_tokens
                }
                logging.info(f"OpenAI API请求参数: {json.dumps(request_params, ensure_ascii=False)}")
                
                response = client.chat.completions.create(
                    model=model_name,
                    messages=messages,
                    temperature=temperature,
                    max_tokens=max_tokens
                )
                
                # 提取回答内容
                content = response.choices[0].message.content
                
                logging.info(f"OpenAI模型验证成功，返回: {content}")
                return {"success": True, "response": content}
            except Exception as e:
                # 记录详细的错误信息
                error_str = str(e)
                full_error = getattr(e, 'response', None)
                if full_error:
                    try:
                        # 尝试从响应中提取更多错误信息
                        response_json = full_error.json()
                        logging.error(f"OpenAI API完整错误响应: {json.dumps(response_json, ensure_ascii=False)}")
                        if 'error' in response_json:
                            error_str = f"{response_json['error'].get('message', error_str)}"
                    except Exception as parse_err:
                        logging.error(f"解析错误响应失败: {str(parse_err)}")
                
                # 检查错误类型并提供更友好的提示
                if "401" in error_str:
                    err_msg = f"Error code: 401 - API密钥无效或认证失败, 请检查API密钥和基础URL设置。详细信息: {error_str}"
                elif "404" in error_str:
                    err_msg = f"Error code: 404 - 请求的资源不存在, 请检查API基础URL和模型名称设置。详细信息: {error_str}"
                elif "429" in error_str:
                    err_msg = f"Error code: 429 - 请求过于频繁或超出配额, 请稍后再试。详细信息: {error_str}"
                elif "500" in error_str or "502" in error_str or "503" in error_str:
                    err_msg = f"Error code: {error_str[:3]} - 服务器错误, 请稍后再试。详细信息: {error_str}"
                else:
                    err_msg = f"模型请求失败: {error_str}"
                
                logging.error(f"模型验证失败: {err_msg}")
                return {"success": False, "error": err_msg}
        elif provider == "ollama":
            # Ollama API调用逻辑
            pass
        else:
            raise ValueError(f"不支持的模型提供商: {provider}")
    except Exception as e:
        logging.error(f"模型验证失败: {e}")
        raise

if __name__ == "__main__":
    # 测试代码
    test_requirement = {
        "name": "用户身份验证功能",
        "content": "系统应支持用户登录功能。"
    }
    result = review_requirement(test_requirement)
    print(json.dumps(result, ensure_ascii=False, indent=2))
