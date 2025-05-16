"""
需求提取模块
负责从需求文档中提取测试需求并进行处理
"""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font

from .document_reader import read_xuqiu_wendang_document, extract_requirement_candidates

def extract_requirements(file_path, requirement_names=None, catalog_file_path=None):
    """
    根据用户确认的需求名称列表提取需求
    参数：
        file_path: Word文档路径
        requirement_names: 用户确认的需求名称列表，如果为None或空列表，则自动提取
        catalog_file_path: 目录文件路径（当需要自动提取需求列表时必需）
    返回：
        测试需求列表（{name: content}）
    """
    # 如果未提供需求名称列表，则自动提取
    if not requirement_names:
        if not catalog_file_path:
            print("[ERROR] 如果未提供需求名称列表，则必须提供目录文件路径")
            return {}
        # 这里获取的是带chapter信息的对象列表
        requirement_candidate_objs = extract_requirement_candidates(file_path, catalog_file_path)
        requirement_names = [obj["name"] for obj in requirement_candidate_objs]
    
    print(f"[DEBUG] 开始提取需求，共{len(requirement_names)}个需求名称")
    for i, name in enumerate(requirement_names):
        print(f"[DEBUG] 需求名称 {i+1}: {name}")
    
    # 然后读取每个需求点的内容
    req_dict = read_xuqiu_wendang_document(file_path, requirement_names)
    print(f"[DEBUG] 提取到{len(req_dict)}个需求内容")
    for name, content in req_dict.items():
        print(f"[DEBUG] 需求「{name}」的内容长度: {len(content)} 字符")
    
    return req_dict



def generate_requirement_excel(requirements, output_path, excel_type="requirement"):
    """
    生成Excel表格 - 支持需求分析表和测试用例表
    
    参数:
        requirements: 测试需求列表
        output_path: 输出文件路径
        excel_type: 表格类型，'requirement'为需求分析表，'test_case'为测试用例表
        
    返回:
        是否成功
    """
    try:
        wb = Workbook()
        ws = wb.active
        ws.title = "需求分析表" if excel_type == "requirement" else "测试用例表"
        
        # 设置表头 - 只保留两列
        headers = ["需求名称", "需求内容"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal='center')
        
        # 填充数据
        for row, (name, content) in enumerate(requirements.items(), 2):
            ws.cell(row=row, column=1, value=name)
            ws.cell(row=row, column=2, value=content)
        
        # 调整列宽
        ws.column_dimensions['A'].width = 30  # 需求名称列的宽度
        ws.column_dimensions['B'].width = 80  # 需求内容列的宽度
        
        # 调整文本自动换行
        for row in range(2, len(requirements) + 2):
            ws.cell(row=row, column=2).alignment = Alignment(wrap_text=True, vertical='top')
        
        # 保存文件
        wb.save(output_path)
        return True
    except Exception as e:
        print(f"生成Excel文件时发生错误: {e}")
        return False
